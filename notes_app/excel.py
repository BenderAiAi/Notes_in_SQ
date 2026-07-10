from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.datetime import from_excel


SOURCE_HEADERS = (
    "Номер",
    "Дата сделки",
    "Дата расчетов",
    "Время",
    "ISIN",
    "Операция",
    "Цена",
    "Кол-во",
    "Объем",
    "Валюта расчетов",
    "Валюта",
    "Комментарий",
    "UID",
)
OUTPUT_HEADERS = (
    "ID",
    "Note",
    "Buy",
    "Amount",
    "Price",
    "Portfolio",
    "Subportfolio",
    "Subaccount",
    "Traded Date Time",
    "Settlement Date",
)
DICTIONARY_HEADERS = (
    "Note",
    "ISIN",
    "Note name SQ",
    "Portfolio",
    "Subportfolio",
    "Subaccount",
)


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("ё", "е").strip()).casefold()


HEADER_LOOKUP = {_normalize_header(header): header for header in SOURCE_HEADERS}


@dataclass
class Issue:
    level: str
    code: str
    message: str
    row: int | None = None

    def as_dict(self) -> dict[str, Any]:
        return {"level": self.level, "code": self.code, "message": self.message, "row": self.row}


@dataclass
class ParsedReport:
    path: Path
    records: list[dict[str, Any]] = field(default_factory=list)
    issues: list[Issue] = field(default_factory=list)
    header_row: int | None = None
    sheet_name: str = ""

    @property
    def errors(self) -> list[Issue]:
        return [issue for issue in self.issues if issue.level == "error"]

    @property
    def warnings(self) -> list[Issue]:
        return [issue for issue in self.issues if issue.level == "warning"]


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except (ValueError, OverflowError):
            return None
    text = str(value or "").strip()
    for pattern in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _parse_time(value: Any) -> time | None:
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    if isinstance(value, (int, float)) and 0 <= value < 1:
        seconds = int(round(float(value) * 86400)) % 86400
        return (datetime.min + timedelta(seconds=seconds)).time()
    text = str(value or "").strip()
    for pattern in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, pattern).time()
        except ValueError:
            continue
    return None


def _parse_decimal(value: Any) -> Decimal | None:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value or "").strip().replace("\u00a0", "").replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _parse_id(value: Any) -> int | str | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    return int(text) if text.isdigit() else text


def _find_header(workbook: Any) -> tuple[Any, int, dict[str, int]] | None:
    for worksheet in workbook.worksheets:
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row), values_only=True), start=1
        ):
            positions: dict[str, int] = {}
            for column_index, value in enumerate(row, start=1):
                canonical = HEADER_LOOKUP.get(_normalize_header(value))
                if canonical:
                    positions[canonical] = column_index
            if len(positions) >= 8 and {"Номер", "Дата сделки", "ISIN", "Операция"}.issubset(positions):
                return worksheet, row_number, positions
    return None


def read_report(path: Path) -> ParsedReport:
    result = ParsedReport(path=path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        result.issues.append(Issue("error", "open_failed", f"Не удалось открыть Excel-файл: {exc}"))
        return result
    try:
        header = _find_header(workbook)
        if not header:
            result.issues.append(Issue("error", "headers_not_found", "Не найдена строка с заголовками исходного отчёта."))
            return result
        worksheet, header_row, positions = header
        result.header_row = header_row
        result.sheet_name = worksheet.title
        missing = [name for name in SOURCE_HEADERS if name not in positions]
        if missing:
            result.issues.append(
                Issue("error", "missing_headers", "Отсутствуют столбцы: " + ", ".join(missing) + ".")
            )
            return result

        for excel_row, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
        ):
            raw = {
                name: values[column - 1] if column - 1 < len(values) else None
                for name, column in positions.items()
            }
            if all(raw.get(name) in (None, "") for name in SOURCE_HEADERS):
                continue
            record = {
                "source_row": excel_row,
                "id": _parse_id(raw["Номер"]),
                "trade_date": _parse_date(raw["Дата сделки"]),
                "settlement_date": _parse_date(raw["Дата расчетов"]),
                "time": _parse_time(raw["Время"]),
                "isin": str(raw["ISIN"] or "").strip().upper(),
                "operation": str(raw["Операция"] or "").strip(),
                "price": _parse_decimal(raw["Цена"]),
                "amount": _parse_decimal(raw["Кол-во"]),
            }
            result.records.append(record)
    finally:
        workbook.close()
    return result


def analyze_report(parsed: ParsedReport, dictionary: dict[str, dict[str, Any]]) -> dict[str, Any]:
    issues = list(parsed.issues)
    if parsed.errors:
        return _analysis_payload(parsed, issues)
    if not parsed.records:
        issues.append(Issue("error", "no_rows", "В отчёте нет строк с данными."))
        return _analysis_payload(parsed, issues)

    seen_ids: dict[Any, int] = {}
    trade_dates_by_isin: dict[str, set[date]] = defaultdict(set)
    settlement_dates_by_isin: dict[str, set[date]] = defaultdict(set)
    for record in parsed.records:
        row = record["source_row"]
        for key, label in (
            ("id", "Номер"), ("trade_date", "Дата сделки"),
            ("settlement_date", "Дата расчетов"), ("time", "Время"),
            ("isin", "ISIN"), ("operation", "Операция"),
            ("price", "Цена"), ("amount", "Кол-во"),
        ):
            if record[key] in (None, ""):
                issues.append(Issue("error", f"empty_{key}", f"Строка {row}: поле «{label}» не заполнено или имеет неверный формат.", row))

        if record["id"] not in (None, ""):
            if record["id"] in seen_ids:
                issues.append(Issue("error", "duplicate_id", f"Строка {row}: номер сделки {record['id']} уже встречался в строке {seen_ids[record['id']] }.", row))
            else:
                seen_ids[record["id"]] = row

        operation = record["operation"].casefold()
        if operation not in {"купля", "продажа"}:
            issues.append(Issue("error", "unknown_operation", f"Строка {row}: неизвестная операция «{record['operation']}». Допустимы «Купля» и «Продажа».", row))

        isin = record["isin"]
        if isin:
            if record["trade_date"]:
                trade_dates_by_isin[isin].add(record["trade_date"])
            if record["settlement_date"]:
                settlement_dates_by_isin[isin].add(record["settlement_date"])
            entry = dictionary.get(isin)
            if not entry:
                issues.append(Issue("error", "missing_isin", f"Строка {row}: ISIN {isin} отсутствует в справочнике.", row))
            else:
                missing_fields = [
                    label for key, label in (
                        ("note_name_sq", "Note name SQ"),
                        ("portfolio", "Portfolio"),
                        ("subportfolio", "Subportfolio"),
                        ("subaccount", "Subaccount"),
                    ) if not str(entry.get(key, "")).strip()
                ]
                if missing_fields:
                    issues.append(Issue("error", "incomplete_dictionary", f"ISIN {isin}: в справочнике не заполнены поля: {', '.join(missing_fields)}."))

        if record["trade_date"] and record["trade_date"] > date.today():
            issues.append(Issue("warning", "future_trade_date", f"Строка {row}: дата сделки {record['trade_date']:%d.%m.%Y} находится в будущем.", row))
        if record["settlement_date"] and record["trade_date"] and record["settlement_date"] < record["trade_date"]:
            issues.append(Issue("warning", "settlement_before_trade", f"Строка {row}: дата расчётов раньше даты сделки.", row))

    for isin, dates in trade_dates_by_isin.items():
        if len(dates) > 1:
            formatted = ", ".join(sorted(item.strftime("%d.%m.%Y") for item in dates))
            issues.append(Issue("error", "multiple_trade_dates", f"ISIN {isin}: найдено несколько дат сделки — {formatted}."))
    for isin, dates in settlement_dates_by_isin.items():
        if len(dates) > 1:
            formatted = ", ".join(sorted(item.strftime("%d.%m.%Y") for item in dates))
            issues.append(Issue("error", "multiple_settlement_dates", f"ISIN {isin}: найдено несколько дат расчётов — {formatted}."))

    return _analysis_payload(parsed, _deduplicate_issues(issues))


def _deduplicate_issues(issues: list[Issue]) -> list[Issue]:
    unique: list[Issue] = []
    seen: set[tuple[str, str, str, int | None]] = set()
    for issue in issues:
        key = (issue.level, issue.code, issue.message, issue.row)
        if key not in seen:
            seen.add(key)
            unique.append(issue)
    return unique


def _analysis_payload(parsed: ParsedReport, issues: list[Issue]) -> dict[str, Any]:
    valid_records = [record for record in parsed.records if record.get("trade_date")]
    trade_dates = sorted({record["trade_date"] for record in valid_records})
    settlement_dates = sorted({record["settlement_date"] for record in parsed.records if record.get("settlement_date")})
    buys = [record for record in parsed.records if record.get("operation", "").casefold() == "купля"]
    sells = [record for record in parsed.records if record.get("operation", "").casefold() == "продажа"]
    def total(records: list[dict[str, Any]]) -> str:
        return str(sum((record.get("amount") or Decimal("0") for record in records), Decimal("0")))
    stat = parsed.path.stat() if parsed.path.exists() else None
    return {
        "path": str(parsed.path),
        "file_name": parsed.path.name,
        "sheet_name": parsed.sheet_name,
        "header_row": parsed.header_row,
        "latest_trade_date": trade_dates[-1].isoformat() if trade_dates else None,
        "trade_dates": [item.isoformat() for item in trade_dates],
        "settlement_dates": [item.isoformat() for item in settlement_dates],
        "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds") if stat else None,
        "summary": {
            "rows": len(parsed.records),
            "buy_trades": len(buys),
            "buy_amount": total(buys),
            "sell_trades": len(sells),
            "sell_amount": total(sells),
            "unique_isin": len({record.get("isin") for record in parsed.records if record.get("isin")}),
        },
        "errors": [issue.as_dict() for issue in issues if issue.level == "error"],
        "warnings": [issue.as_dict() for issue in issues if issue.level == "warning"],
        "can_generate": not any(issue.level == "error" for issue in issues),
    }


def generate_output(parsed: ParsedReport, dictionary: dict[str, dict[str, Any]], output_dir: Path) -> Path:
    analysis = analyze_report(parsed, dictionary)
    if not analysis["can_generate"]:
        raise ValueError("Отчёт содержит блокирующие ошибки.")
    report_date = max(record["trade_date"] for record in parsed.records if record["trade_date"])
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"Note_trades_sq_{report_date:%d%m%Y}.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "note_trades"
    worksheet.append(OUTPUT_HEADERS)
    for record in parsed.records:
        entry = dictionary[record["isin"]]
        traded = datetime.combine(record["trade_date"], record["time"])
        settlement = datetime.combine(record["settlement_date"], record["time"])
        worksheet.append(
            (
                record["id"],
                entry["note_name_sq"],
                "B" if record["operation"].casefold() == "купля" else "S",
                float(record["amount"]),
                float(record["price"]),
                entry["portfolio"],
                entry["subportfolio"],
                entry["subaccount"],
                traded,
                settlement,
            )
        )

    header_fill = PatternFill("solid", fgColor="173B3F")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 24
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = (15, 38, 10, 14, 14, 24, 34, 28, 23, 23)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    for row in worksheet.iter_rows(min_row=2):
        row[3].number_format = "0.############"
        row[4].number_format = "0.############"
        row[8].number_format = "dd.mm.yyyy hh:mm:ss"
        row[9].number_format = "dd.mm.yyyy hh:mm:ss"
    workbook.save(output_path)
    return output_path


def read_dictionary_workbook(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header_positions: dict[str, int] = {}
        expected = {_normalize_header(value): value for value in DICTIONARY_HEADERS}
        header_row = None
        for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row), values_only=True), start=1):
            found = {expected[_normalize_header(value)]: index for index, value in enumerate(row) if _normalize_header(value) in expected}
            if "ISIN" in found and len(found) >= 4:
                header_positions = found
                header_row = row_number
                break
        if header_row is None:
            raise ValueError("В файле справочника не найдена шапка с полем ISIN.")
        missing = [header for header in DICTIONARY_HEADERS if header not in header_positions]
        if missing:
            raise ValueError("В справочнике отсутствуют столбцы: " + ", ".join(missing) + ".")
        field_by_header = {
            "Note": "note", "ISIN": "isin", "Note name SQ": "note_name_sq",
            "Portfolio": "portfolio", "Subportfolio": "subportfolio", "Subaccount": "subaccount",
        }
        entries = []
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            entry = {
                field_by_header[header]: str(row[index] or "").strip() if index < len(row) else ""
                for header, index in header_positions.items()
            }
            if entry["isin"]:
                entries.append(entry)
        return entries
    finally:
        workbook.close()


def export_dictionary_workbook(entries: list[dict[str, Any]], path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "dictionary"
    worksheet.append(DICTIONARY_HEADERS)
    for entry in entries:
        worksheet.append((entry["note"], entry["isin"], entry["note_name_sq"], entry["portfolio"], entry["subportfolio"], entry["subaccount"]))
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor="173B3F")
        cell.font = Font(color="FFFFFF", bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column, width in zip("ABCDEF", (28, 22, 42, 28, 38, 32)):
        worksheet.column_dimensions[column].width = width
    workbook.save(path)
