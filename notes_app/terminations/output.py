from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .report_builder import (
    DEFAULT_CY_FIELD,
    build_dubai_report,
    build_trs_report,
)


HISTORY_EXPORT_HEADERS = (
    "Дата расторжения",
    "Номер контракта",
    "Тип",
    "Ставка, %",
    "Сумма расторжения",
    "Валюта",
    "Номинал",
    "Валюта номинала",
    "Тип расторжения",
    "Кипр",
    "Файл-источник",
    "Записано",
)

CONTRACT_TYPE_LABELS = {
    "dubai": "Dubai",
    "trs": "TRS",
    "both": "Dubai+TRS",
    "not_found": "не найден",
}
DISSOLUTION_LABELS = {"manual": "Ручное", "mssp": "МССП", "unknown": "—"}


def output_file_names(today_date: str) -> dict[str, str]:
    return {
        "dubai": f"{today_date}-Dubai_term.xlsx",
        "trs": f"{today_date}-TRS_term.xlsx",
    }


def existing_outputs(today_date: str, output_dir: Path) -> list[str]:
    names = output_file_names(today_date)
    return [name for name in names.values() if (output_dir / name).exists()]


def generate_reports(
    clean_df: pd.DataFrame,
    mongo_df_1: pd.DataFrame,
    mongo_df_2: pd.DataFrame,
    today_date: str,
    output_dir: Path,
    cy_field: str = DEFAULT_CY_FIELD,
) -> dict[str, Any]:
    dubai = build_dubai_report(mongo_df_1, clean_df, today_date, cy_field)
    trs = build_trs_report(mongo_df_2, clean_df, today_date)
    output_dir.mkdir(parents=True, exist_ok=True)
    names = output_file_names(today_date)

    saved: dict[str, Any] = {
        "dubai_path": None,
        "trs_path": None,
        "dubai_count": int(len(dubai)),
        "trs_count": int(len(trs)),
    }
    if not dubai.empty:
        dubai_path = output_dir / names["dubai"]
        with pd.ExcelWriter(dubai_path, datetime_format="dd.mm.yyyy") as writer:
            dubai.to_excel(writer, sheet_name="dubai_dissolution", index=False)
        saved["dubai_path"] = str(dubai_path)
    if not trs.empty:
        trs_path = output_dir / names["trs"]
        with pd.ExcelWriter(trs_path, datetime_format="dd.mm.yyyy") as writer:
            trs.to_excel(writer, sheet_name="second_db_dissolution", index=False)
        saved["trs_path"] = str(trs_path)
    return saved


def export_history_workbook(rows: list[dict[str, Any]], path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "history"
    worksheet.append(HISTORY_EXPORT_HEADERS)
    for row in rows:
        worksheet.append(
            (
                _fmt_date(row.get("termination_date")),
                row.get("contract_number", ""),
                CONTRACT_TYPE_LABELS.get(row.get("contract_type", ""), row.get("contract_type", "")),
                row.get("rate_pct"),
                row.get("termination_amount"),
                row.get("return_currency", ""),
                row.get("notional"),
                row.get("notional_currency", ""),
                DISSOLUTION_LABELS.get(row.get("dissolution_type", ""), row.get("dissolution_type", "")),
                "Да" if row.get("cyprus_flag") else "",
                row.get("source_file", ""),
                row.get("updated_at", ""),
            )
        )
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor="173B3F")
        cell.font = Font(color="FFFFFF", bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = (16, 24, 12, 12, 18, 10, 18, 14, 14, 8, 26, 20)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    for report_row in worksheet.iter_rows(min_row=2):
        report_row[3].number_format = "0.0000"
        report_row[4].number_format = "# ##0.00"
        report_row[6].number_format = "# ##0.00"
    workbook.save(path)
    return path


def _fmt_date(iso: str | None) -> str:
    if not iso:
        return ""
    parts = str(iso)[:10].split("-")
    if len(parts) == 3:
        return f"{parts[2]}.{parts[1]}.{parts[0]}"
    return str(iso)
