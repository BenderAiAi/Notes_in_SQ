from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from notes_app.excel import OUTPUT_HEADERS, SOURCE_HEADERS, analyze_report, generate_output, read_report


def make_source(path: Path, rows: list[list[object]]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Сделки"
    # Пустая первая ячейка имитирует техническую нумерацию строк из исходного отчёта.
    sheet.append([None, *SOURCE_HEADERS])
    for index, row in enumerate(rows, start=1):
        sheet.append([index, *row])
    workbook.save(path)
    return path


def dictionary() -> dict[str, dict[str, str]]:
    return {
        "XS2777123369": {
            "note": "Reference note",
            "isin": "XS2777123369",
            "note_name_sq": "series 111 isin:XS2777123369",
            "portfolio": "EMTN BRV SP",
            "subportfolio": "Phoenixes with unprotected puts",
            "subaccount": "777: Phoenix Notes",
        }
    }


def valid_rows() -> list[list[object]]:
    return [
        ["2 991 432", "10.07.2026", "13.07.2026", "10:37:14", "XS2777123369", "Купля", "41,01", "35", "1 323,57", "SUR", "USD", "D99/777", "249 235"],
        ["2 991 433", "10.07.2026", "13.07.2026", "11:01:57", "XS2777123369", "Продажа", "41,01", "9", "3 312,55", "SUR", "USD", "Aq/777", "249 235"],
    ]


def test_read_analyze_and_generate_exact_output(tmp_path: Path) -> None:
    source = make_source(tmp_path / "Ежедневный отчёт 10.07.2026.xlsx", valid_rows())
    parsed = read_report(source)
    analysis = analyze_report(parsed, dictionary())

    assert analysis["can_generate"] is True
    assert analysis["summary"] == {
        "rows": 2,
        "buy_trades": 1,
        "buy_amount": "35",
        "sell_trades": 1,
        "sell_amount": "9",
        "unique_isin": 1,
    }
    assert analysis["latest_trade_date"] == "2026-07-10"

    output = generate_output(parsed, dictionary(), tmp_path / "ready")
    assert output.name == "Note_trades_sq_10072026.xlsx"
    workbook = load_workbook(output, data_only=True)
    try:
        assert workbook.sheetnames == ["note_trades"]
        sheet = workbook["note_trades"]
        assert tuple(cell.value for cell in sheet[1]) == OUTPUT_HEADERS
        assert sheet["A2"].value == 2991432
        assert sheet["B2"].value == "series 111 isin:XS2777123369"
        assert sheet["C2"].value == "B"
        assert sheet["C3"].value == "S"
        assert sheet["D2"].value == 35
        assert sheet["E2"].value == 41.01
        assert sheet["I2"].value == datetime(2026, 7, 10, 10, 37, 14)
        assert sheet["J2"].value == datetime(2026, 7, 13, 10, 37, 14)
    finally:
        workbook.close()


def test_missing_dictionary_entry_is_explained(tmp_path: Path) -> None:
    source = make_source(tmp_path / "report.xlsx", valid_rows())
    analysis = analyze_report(read_report(source), {})
    messages = [issue["message"] for issue in analysis["errors"]]
    assert analysis["can_generate"] is False
    assert any("ISIN XS2777123369 отсутствует в справочнике" in message for message in messages)


def test_multiple_dates_are_checked_per_isin(tmp_path: Path) -> None:
    rows = valid_rows()
    rows[1][1] = "11.07.2026"
    source = make_source(tmp_path / "report.xlsx", rows)
    analysis = analyze_report(read_report(source), dictionary())
    codes = {issue["code"] for issue in analysis["errors"]}
    assert "multiple_trade_dates" in codes


def test_duplicate_trade_id_has_source_row(tmp_path: Path) -> None:
    rows = valid_rows()
    rows[1][0] = rows[0][0]
    source = make_source(tmp_path / "report.xlsx", rows)
    analysis = analyze_report(read_report(source), dictionary())
    duplicate = next(issue for issue in analysis["errors"] if issue["code"] == "duplicate_id")
    assert duplicate["row"] == 3
    assert "строке 2" in duplicate["message"]


def test_optional_summary_in_ninth_column_is_not_a_trade(tmp_path: Path) -> None:
    source = make_source(tmp_path / "report_with_summary.xlsx", valid_rows())
    workbook = load_workbook(source)
    try:
        sheet = workbook.active
        # Первый столбец технический, поэтому «Объем» находится в J.
        # В реальном отчёте эта необязательная строка идёт сразу после сделок.
        sheet.cell(row=4, column=10, value="Total: 4 636,12")
        # Дополнительные варианты служебного хвоста также не должны становиться
        # сделками, если значение оказалось в другом одиночном столбце.
        sheet.cell(row=5, column=9, value=44)
        sheet.cell(row=6, column=7, value="Итого")
        workbook.save(source)
    finally:
        workbook.close()

    parsed = read_report(source)
    analysis = analyze_report(parsed, dictionary())

    assert len(parsed.records) == 2
    assert analysis["summary"]["rows"] == 2
    assert analysis["can_generate"] is True
    assert analysis["warnings"] == []
    assert len(analysis["notices"]) == 3
    assert "J4 (Объем) = «Total: 4 636,12»" in analysis["notices"][0]["message"]


def test_amount_with_trailing_comma_is_written_as_integer_number(tmp_path: Path) -> None:
    rows = valid_rows()
    rows[0][7] = "4,"
    rows[1][7] = "17,"
    source = make_source(tmp_path / "integer_amounts.xlsx", rows)
    parsed = read_report(source)
    analysis = analyze_report(parsed, dictionary())

    assert analysis["can_generate"] is True
    assert [record["amount"] for record in parsed.records] == [4, 17]

    output = generate_output(parsed, dictionary(), tmp_path / "ready")
    workbook = load_workbook(output, data_only=True)
    try:
        sheet = workbook["note_trades"]
        assert sheet["D2"].value == 4
        assert sheet["D3"].value == 17
        assert sheet["D2"].data_type == "n"
        assert sheet["D3"].data_type == "n"
        assert sheet["D2"].number_format == "0"
    finally:
        workbook.close()


def test_fractional_amount_is_rejected(tmp_path: Path) -> None:
    rows = valid_rows()
    rows[0][7] = "4,5"
    source = make_source(tmp_path / "fractional_amount.xlsx", rows)
    analysis = analyze_report(read_report(source), dictionary())

    error = next(issue for issue in analysis["errors"] if issue["code"] == "fractional_amount")
    assert analysis["can_generate"] is False
    assert "Amount должен быть целым числом" in error["message"]
